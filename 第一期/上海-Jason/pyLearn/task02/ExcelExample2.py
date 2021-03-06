# _*_ coding : utf-8 _*_
__author__ = "Jason"

'''
读取已存在的Excel文件
'''

from openpyxl import load_workbook
if __name__ == "__main__":
    print("读取已存在的Excel文件")

    wb = load_workbook("简单Excel写示例.xlsx")

    #获取所有sheet名，返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))#<class 'list'>

    #遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作簿名称：",sh)

    #获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    #读取sheet A1、A2、B12、2单元格内容

    #读取A1单元格的值
    A1 = ws["A1"].value
    print("A1单元格的值：",A1)

    #读取A2、B2、C2
    for index in ("A2","B2","C2"):
        print(index,"单元格的值：",ws[index].value)

    #读取空值的单元格，openpyxl对于空值的单元格，返回None
    F1 = ws["F1"].value
    print("F1单元格的值：",F1)